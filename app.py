import openpyxl as xl
from openpyxl.chart import BarChart,Reference
#from openpyxl Package's chart module import classes BarChart and Reference


def processing(filename):  #Reusable function to process many spreadsheets
    workbook = xl.load_workbook(filename)
    sheet=workbook['Sheet1']

    for row in range(2,sheet.max_row+1): #from 2nd row to last row+1 else last row excluded
        cell3=sheet.cell(row,3)
        corrected_price=cell3.value*0.9
        corrected_price_cell=sheet.cell(row,4)   #All corrected values in new column
        corrected_price_cell.value = corrected_price #setting value of the cell to corrected price

    values=Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4) #passed all values into the constructor
    charts=BarChart() #created an object of bar chart
    charts.add_data(values) #adding values to the bar chart
    sheet.add_chart(charts,'F2') #adding to bar chart to the 5th column

    workbook.save(filename) #save in a new file to avoid accidentally overwrite the new file